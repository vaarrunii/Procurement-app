# core/views.py

import csv
import os
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST, require_GET
from datetime import date
from decimal import Decimal
import datetime

# For analytics
from django.db.models import Sum, F, Q, Count # Added Count

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import (
    Supplier, Category, PurchaseOrder, Invoice, SpendEntry,
    SupplierProductPricing, SupplierContract, SupplierDiscount, AlternateSupplier
)
from django.db.models import ForeignKey, OneToOneField
from django.db.models.fields.related import ManyToManyField
from django.apps import apps # Import apps to dynamically get models

@login_required
def data_home(request):
    return render(request, 'core/data_home.html')

@login_required
def summary_page(request):
    """
    Dedicated view for comprehensive analytics across models.
    """
    # --- Analytics Logic for Summary Page ---
    # Invoice Analytics
    total_invoices_summary = Invoice.objects.count()
    total_pending_invoices_summary = Invoice.objects.filter(Q(status__iexact='Pending')).count()
    sum_pending_invoice_amount_summary = Invoice.objects.filter(Q(status__iexact='Pending')).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
    total_paid_invoices_summary = Invoice.objects.filter(Q(status__iexact='Paid')).count()
    sum_paid_invoice_amount_summary = Invoice.objects.filter(Q(status__iexact='Paid')).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')

    # Purchase Order Analytics
    total_purchase_orders_summary = PurchaseOrder.objects.count()
    total_pending_pos_summary = PurchaseOrder.objects.filter(Q(status__iexact='Pending')).count()
    sum_pending_po_amount_summary = PurchaseOrder.objects.filter(Q(status__iexact='Pending')).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
    total_approved_pos_summary = PurchaseOrder.objects.filter(Q(status__iexact='Approved')).count()
    sum_approved_po_amount_summary = PurchaseOrder.objects.filter(Q(status__iexact='Approved')).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')

    # Spend Entry Analytics
    total_spend_entries_summary = SpendEntry.objects.count()
    total_spend_amount_summary = SpendEntry.objects.aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
    spend_by_category_summary = list(SpendEntry.objects.values('category__name').annotate(total=Sum('amount')).order_by('-total'))
    spend_by_supplier_summary = list(SpendEntry.objects.values('supplier__name').annotate(total=Sum('amount')).order_by('-total'))

    # Supplier Analytics
    total_suppliers_summary = Supplier.objects.count()
    active_suppliers_summary = Supplier.objects.filter(is_active=True).count()
    inactive_suppliers_summary = Supplier.objects.filter(is_active=False).count()
    suppliers_by_type_summary = list(Supplier.objects.values('type').annotate(count=Count('id')).order_by('-count'))


    context = {
        'total_invoices_summary': total_invoices_summary,
        'total_pending_invoices_summary': total_pending_invoices_summary,
        'sum_pending_invoice_amount_summary': sum_pending_invoice_amount_summary,
        'total_paid_invoices_summary': total_paid_invoices_summary,
        'sum_paid_invoice_amount_summary': sum_paid_invoice_amount_summary,

        'total_purchase_orders_summary': total_purchase_orders_summary,
        'total_pending_pos_summary': total_pending_pos_summary,
        'sum_pending_po_amount_summary': sum_pending_po_amount_summary,
        'total_approved_pos_summary': total_approved_pos_summary,
        'sum_approved_po_amount_summary': sum_approved_po_amount_summary,

        'total_spend_entries_summary': total_spend_entries_summary,
        'total_spend_amount_summary': total_spend_amount_summary,
        'spend_by_category_summary': spend_by_category_summary,
        'spend_by_supplier_summary': spend_by_supplier_summary,

        'total_suppliers_summary': total_suppliers_summary,
        'active_suppliers_summary': active_suppliers_summary,
        'inactive_suppliers_summary': inactive_suppliers_summary,
        'suppliers_by_type_summary': suppliers_by_type_summary,
    }
    return render(request, 'core/summary_page.html', context)


def _render_model_list(request, model, title, fields_to_display):
    """
    Helper function to fetch all objects of a model and render them.
    Prepares data and field names for DataTables consumption via JSON.
    Also calculates and passes basic analytics for the current model.
    """
    print(f"\n--- Entering _render_model_list for: {title} ---")
    print(f"Model: {model.__name__}, Fields to Display (from view func): {fields_to_display}")

    objects_data = [] # Initialize in case of error
    try:
        # Ensure 'id' is always included for actions even if not in fields_to_display
        all_fields_for_values = list(set(fields_to_display + ['id']))
        objects_data = list(model.objects.all().values(*all_fields_for_values))
        print(f"Step 1: Fetched {len(objects_data)} records from {model.__name__}.")
        if objects_data:
            print(f"Step 1: First record's data keys (from .values()): {list(objects_data[0].keys())}")
            print(f"Step 1: First record's full data (from .values()): {objects_data[0]}")
        else:
            print(f"Step 1: No records fetched from the database for {model.__name__}.")
    except Exception as e:
        print(f"Step 1: ERROR fetching data for {model.__name__}: {e}")
        objects_data = [] # Ensure it's an empty list on error for json_script

    data_list_for_json = [] # Data after cleaning for JSON
    for obj in objects_data:
        cleaned_obj = {}
        for key, value in obj.items():
            if isinstance(value, float):
                cleaned_obj[key] = round(value, 2)
            elif isinstance(value, Decimal):
                cleaned_obj[key] = str(value) # Convert Decimal to string
            elif isinstance(value, date):
                cleaned_obj[key] = value.isoformat() # Convert date to ISO format string
            elif isinstance(value, datetime.datetime):
                cleaned_obj[key] = value.isoformat()
            else:
                cleaned_obj[key] = value
        data_list_for_json.append(cleaned_obj)

    print(f"Step 2: Prepared data_list_for_json length (after cleaning): {len(data_list_for_json)}")
    if data_list_for_json:
        print(f"Step 2: First cleaned record example: {data_list_for_json[0]}")
    else:
        print("Step 2: Cleaned data_list_for_json is empty.")

    # Convert model name to URL-friendly format (e.g., 'Supplier' -> 'supplier-data')
    url_model_name = model.__name__.lower().replace('_', '-') + '-data'

    # --- Analytics for the current model (for display on model_list page) ---
    analytics_data = {}
    if model == Invoice:
        analytics_data['total_records'] = Invoice.objects.count()
        analytics_data['total_pending'] = Invoice.objects.filter(Q(status__iexact='Pending')).count()
        analytics_data['sum_pending_amount'] = Invoice.objects.filter(Q(status__iexact='Pending')).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        analytics_data['total_paid'] = Invoice.objects.filter(Q(status__iexact='Paid')).count()
        analytics_data['sum_paid_amount'] = Invoice.objects.filter(Q(status__iexact='Paid')).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        analytics_data['analytics_title'] = "Invoice Status Summary"
    elif model == PurchaseOrder:
        analytics_data['total_records'] = PurchaseOrder.objects.count()
        analytics_data['total_pending'] = PurchaseOrder.objects.filter(Q(status__iexact='Pending')).count()
        analytics_data['sum_pending_amount'] = PurchaseOrder.objects.filter(Q(status__iexact='Pending')).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        analytics_data['total_approved'] = PurchaseOrder.objects.filter(Q(status__iexact='Approved')).count()
        analytics_data['sum_approved_amount'] = PurchaseOrder.objects.filter(Q(status__iexact='Approved')).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        analytics_data['analytics_title'] = "Purchase Order Status Summary"
    elif model == SpendEntry:
        analytics_data['total_records'] = SpendEntry.objects.count()
        analytics_data['total_spend_amount'] = SpendEntry.objects.aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        analytics_data['analytics_title'] = "Spend Entry Summary"
    elif model == Supplier:
        analytics_data['total_records'] = Supplier.objects.count()
        analytics_data['active_suppliers'] = Supplier.objects.filter(is_active=True).count()
        analytics_data['inactive_suppliers'] = Supplier.objects.filter(is_active=False).count()
        analytics_data['analytics_title'] = "Supplier Status Summary"
    else:
        analytics_data['total_records'] = model.objects.count()
        analytics_data['analytics_title'] = f"{model.__name__} Data Summary"

    context = {
        'model_name': title,
        'data': data_list_for_json,
        'fields': fields_to_display, # Pass original fields for header generation
        'url_model_name': url_model_name, # Pass the URL-friendly model name for JS to build links
        'analytics': analytics_data, # NEW: Pass analytics data to template
    }
    print(f"Step 3: Context 'data' length: {len(context['data'])}, 'fields' length: {len(context['fields'])}, 'url_model_name': {context['url_model_name']}")
    print(f"Analytics Data (for {model.__name__}): {analytics_data}") # DEBUG
    print(f"--- Exiting _render_model_list for: {title} ---\n")
    return render(request, 'core/model_list.html', context)

@login_required
def export_model_excel(request, model_name):
    # This function remains unchanged
    model_map = {
        'supplier-data': Supplier,
        'category-data': Category,
        'purchase-order-data': PurchaseOrder,
        'invoice-data': Invoice,
        'spend-entry-data': SpendEntry,
        'supplier-product-pricing-data': SupplierProductPricing,
        'supplier-contract-data': SupplierContract,
        'supplier-discount-data': SupplierDiscount,
        'alternate-supplier-data': AlternateSupplier,
    }

    model = model_map.get(model_name)
    if not model:
        return HttpResponse("Model not found.", status=404)

    fields_to_export = []
    for field in model._meta.get_fields(include_hidden=False):
        if field.concrete and not field.auto_created and not isinstance(field, ManyToManyField):
            fields_to_export.append(field)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = model.__name__ + " Data"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border_thin = Side(style='thin', color="000000")
    header_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    headers = []
    for field in fields_to_export:
        headers.append(field.verbose_name.replace('_', ' ').title())

    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for obj in model.objects.all():
        row_data = []
        for field in fields_to_export:
            value = getattr(obj, field.name)

            if isinstance(field, (ForeignKey, OneToOneField)):
                row_data.append(str(value) if value else 'N/A')
            elif isinstance(value, (date, datetime.date, datetime.datetime)):
                row_data.append(value)
            elif isinstance(value, Decimal):
                row_data.append(float(value))
            elif isinstance(value, bool):
                row_data.append("Yes" if value else "No")
            else:
                row_data.append(value)
        worksheet.append(row_data)

    for col_idx, col in enumerate(worksheet.columns, 1):
        max_length = 0
        column = get_column_letter(col_idx)
        for cell in col:
            try:
                cell_value_str = str(cell.value) if cell.value is not None else ""
                if len(cell_value_str) > max_length:
                    max_length = len(cell_value_str)
            except:
                pass
        adjusted_width = (max_length + 2)
        if adjusted_width > 75:
            adjusted_width = 75
        elif adjusted_width < 10:
            adjusted_width = 10
        worksheet.column_dimensions[column].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{model.__name__.lower()}_data.xlsx"'
    workbook.save(response)
    return response

# Helper function to get model from string name (used by edit/delete views)
def _get_model_from_name(model_name):
    # Define a consistent mapping from URL names to actual Django Model classes
    model_map = {
        'supplier-data': Supplier,
        'category-data': Category,
        'purchase-order-data': PurchaseOrder,
        'invoice-data': Invoice,
        'spend-entry-data': SpendEntry,
        'supplier-product-pricing-data': SupplierProductPricing,
        'supplier-contract-data': SupplierContract,
        'supplier-discount-data': SupplierDiscount,
        'alternate-supplier-data': AlternateSupplier,
    }
    return model_map.get(model_name)

@login_required
@require_GET
def edit_model_record(request, model_name, pk):
    model = _get_model_from_name(model_name)
    if not model:
        return HttpResponse("Model not found.", status=404)

    instance = get_object_or_404(model, pk=pk)

    # --- IMPORTANT: Placeholder for Actual Edit Form ---
    context = {
        'model_name': model.__name__,
        'record_id': pk,
        'instance': instance, # You can pass the instance to display its details
    }
    return render(request, 'core/edit_record.html', context)

@login_required
@require_POST
def delete_model_record(request, model_name, pk):
    model = _get_model_from_name(model_name)
    if not model:
        return JsonResponse({'status': 'error', 'message': 'Model not found.'}, status=404)

    try:
        instance = get_object_or_404(model, pk=pk)
        instance.delete()
        return JsonResponse({'status': 'success', 'message': 'Record deleted successfully.'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


# Specific views for each model (these already provide the correct field names including '_id')
@login_required
def supplier_list(request):
    fields = ['id', 'name', 'contact_email', 'phone', 'address', 'gstin', 'pan', 'score', 'type', 'is_active', 'share_of_business', 'lead_time_days', 'base_currency', 'unit_of_measure']
    return _render_model_list(request, Supplier, 'Supplier Data', fields)

@login_required
def category_list(request):
    fields = ['id', 'name', 'parent_id']
    return _render_model_list(request, Category, 'Category Data', fields)

@login_required
def purchase_order_list(request):
    fields = ['id', 'po_number', 'supplier_id', 'category_id', 'amount', 'issue_date', 'status']
    return _render_model_list(request, PurchaseOrder, 'Purchase Order Data', fields)

@login_required
def invoice_list(request):
    fields = ['id', 'invoice_number', 'supplier_id', 'purchase_order_id', 'invoice_date', 'due_date', 'paid_date', 'amount', 'status']
    return _render_model_list(request, Invoice, 'Invoice Data', fields)

@login_required
def spend_entry_list(request):
    fields = ['id', 'category_id', 'supplier_id', 'date', 'amount', 'cost_center', 'description']
    return _render_model_list(request, SpendEntry, 'Spend Entry Data', fields)

@login_required
def supplier_product_pricing_list(request):
    fields = ['id', 'supplier_id', 'product_name', 'price', 'currency', 'unit_of_measure']
    return _render_model_list(request, SupplierProductPricing, 'Supplier Product Pricing Data', fields)

@login_required
def supplier_contract_list(request):
    fields = ['id', 'supplier_id', 'contract_name', 'start_date', 'end_date', 'terms']
    return _render_model_list(request, SupplierContract, 'Supplier Contract Data', fields)

@login_required
def supplier_discount_list(request):
    fields = ['id', 'supplier_id', 'product_name', 'discount_percent', 'valid_from', 'valid_to']
    return _render_model_list(request, SupplierDiscount, 'Alternate Supplier Data', fields)

@login_required
def alternate_supplier_list(request):
    fields = ['id', 'product_name', 'primary_supplier_id', 'alternate_supplier_id', 'lead_time_days']
    return _render_model_list(request, AlternateSupplier, 'Alternate Supplier Data', fields)